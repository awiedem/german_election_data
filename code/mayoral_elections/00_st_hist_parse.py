#!/usr/bin/env python3
"""Stage-0 parser for the Sachsen-Anhalt (ST) HISTORICAL Bürgermeisterwahlen file.

Source: Statistisches Landesamt Sachsen-Anhalt
  data/mayoral_elections/raw/sachsen_anhalt/2026_0661_BM-Wahl_ab_1994.xlsx
  (sheets: "Erläuterung" = spec/caveats, "BM-Wahlen Sachsen-Anhalt" = data;
   title row states "Bürgermeisterwahlen in Sachsen-Anhalt ab 1994 (Stand: ...)")

This SUPERSEDES the bmbm.csv snapshot as the primary ST source: it covers
1994-2026 (4142 elections, 2011 historical AGS → 218 current Gemeinden) where
bmbm.csv held only the most-recent election per Gemeinde (2019-2026, 218 rows).

Layout — one wide row per (Gemeinde, Wahltag). 3 header rows (r3/r4/r5), data
from r6. 25 metadata columns then 26 candidate blocks B01..B26 of 7 columns:

  [ 0] AGS* am Wahltag      [ 1] Name (am Wahltag)
  [ 2] AGS aktuell          [ 3] Name (aktuell)      <- StaLA's own crosswalk
  [ 4] Ergebnisart (E=endgültig, V=vorläufig)
  [ 5] Wahldatum Hauptwahl  [ 6] Wahldatum Stichwahl
  [ 7] Amtsantrittsdatum    [ 8] Titel   [ 9] Name   [10] Vorname   (Wahlsieger)
  [11] Art des Amtes        [12] Partei Kurzbez.     [13] Partei Bezeichnung
  [14] Geschlecht           [15] Geburtsjahr
  [16..19] Hauptwahl  Wahlberechtigte/Wähler/Ungültige/Gültige
  [20..23] Stichwahl  Wahlberechtigte/Wähler/Ungültige/Gültige
  [24] Anzahl Wahlvorschläge
  [25 + 7k .. 31 + 7k] B{k+1}: Titel, Name, Vorname, Partei, Geschlecht,
                                Stimmen(HW), Stimmen(SW)

Quirks / traps (all handled below — each was observed in the file):
  * SENTINELS (per the sheet's own Zeichenerklärung):
      '-' = "nichts vorhanden (genau Null)"   -> not applicable / no such round
      '.' = "Zahlenwert unbekannt"            -> unknown -> NA
    Both must map to NA for numerics; '.' in a NAME column means the candidate
    exists but is un-named (500 slots), which is NOT the same as an empty slot.
  * MIXED DATE FORMATS in the Stichwahl column: 403 rows are ISO datetimes but
    84 are GERMAN "dd.mm.yyyy" strings (e.g. Leuna 1994 "26.06.1994"). Parsing
    only ISO would silently drop 84 runoffs. Amtsantritt additionally has one
    month-only value ("12.2021") which is treated as unknown.
  * OFFICE CLASSIFIER — "Art des Amtes": 'HaOB' -> Oberbürgermeisterwahl;
    'Ha' (hauptamtlich), '-'/'' (ehrenamtlich) -> Bürgermeisterwahl. There is
    ONE lowercase 'ha' typo (Seeland, Stadt 2016) — matched case-insensitively.
  * 1994 IS SPARSE by design: the Erläuterung notes that for many 1994 elections
    only the winner is known (599/1299 rows lack Gültige Stimmen, 756 lack
    Wahlberechtigte, 499 candidate slots are un-named). Later years are ~complete.
  * The final sheet row is a FOOTNOTE ("*Kreisgebietsreformen: 01. Juli 1994,
    01. Juli 2007"), not data — rows are kept only if col 0 is an 8-digit AGS.
  * `ags` is emitted as AGS *am Wahltag* (original boundaries — what the harm
    stage crosswalks by (ags, election_year)). `ags_current` carries StaLA's own
    mapping to today's 218 Gemeinden and is kept for reference/fallback.

Known SOURCE anomalies (reported, not silently patched):
  * Halle (Saale) 2025-02-02: candidate votes sum to 87407 vs Gültige 87437 (-30).
  * Leuna, Stadt 1994-06-12: candidate votes sum to 4010 vs Gültige 3910 (+100).

Merge: the historical file is missing a handful of elections that the older
bmbm.csv snapshot does carry. Those are appended verbatim from
`st_bmbm_parsed.csv` (written by 00_st_stala_parse.py) for any (ags,
election_date) the historical file lacks — so this output is a strict superset
of the previous one and no coverage can regress.

Output (candidate-level long, one row per candidate per round — same schema as
before, so stages 01/01b consume it unchanged):
  data/mayoral_elections/raw/sachsen_anhalt/st_stala_parsed.csv

Run order:
  python3 code/mayoral_elections/00_st_stala_parse.py   # -> st_bmbm_parsed.csv
  python3 code/mayoral_elections/00_st_hist_parse.py    # -> st_stala_parsed.csv
"""

import csv
import os
import re
import sys
import unicodedata
from typing import Optional

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
RAW_DIR = os.path.join(ROOT, "data", "mayoral_elections", "raw", "sachsen_anhalt")
IN_XLSX = os.path.join(RAW_DIR, "2026_0661_BM-Wahl_ab_1994.xlsx")
IN_BMBM = os.path.join(RAW_DIR, "st_bmbm_parsed.csv")
OUT_CSV = os.path.join(RAW_DIR, "st_stala_parsed.csv")

SHEET = "BM-Wahlen Sachsen-Anhalt"
STATE = "15"
STATE_NAME = "Sachsen-Anhalt"
N_CAND_SLOTS = 26
FIRST_B_COL = 25
B_WIDTH = 7
SOURCE_TAG = "StaLA 2026_0661_BM-Wahl_ab_1994.xlsx"
BMBM_TAG = "StaLA bmbm.csv (Dezernat 13)"

# Sentinels used by the sheet's Zeichenerklärung.
NULLISH = {"", "-", "."}

ISO_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})")
GER_RE = re.compile(r"^(\d{2})\.(\d{2})\.(\d{4})$")

# Columns
C_AGS_WT, C_NAME_WT, C_AGS_AK, C_NAME_AK = 0, 1, 2, 3
C_ERGART = 4
C_HW_DATE, C_SW_DATE, C_AMTSANTRITT = 5, 6, 7
C_W_TITEL, C_W_NAME, C_W_VORNAME = 8, 9, 10
C_ART, C_W_PARTEI_K, C_W_PARTEI_B, C_W_GESCHL, C_W_GEBJAHR = 11, 12, 13, 14, 15
C_HW_WBER, C_HW_WAEHL, C_HW_UNG, C_HW_GUELT = 16, 17, 18, 19
C_SW_WBER, C_SW_WAEHL, C_SW_UNG, C_SW_GUELT = 20, 21, 22, 23
C_ANZ = 24


def cell(row, i) -> str:
    v = row[i] if i < len(row) else None
    return "" if v is None else str(v).strip()


def clean_text(s: str) -> str:
    """Sentinel-aware text: '-' and '.' both mean 'no value here'."""
    s = (s or "").replace(" ", " ").strip()
    return "" if s in NULLISH else s


def clean_num(s: str) -> Optional[float]:
    """German integer with optional space/nbsp thousands separators.

    '-' and '.' are the sheet's sentinels and both yield None (NA).
    """
    s = (s or "").replace(" ", " ").replace(" ", " ").strip()
    if s in NULLISH:
        return None
    s = s.replace(" ", "").replace(".", "") if s.count(".") > 1 else s.replace(" ", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(s: str) -> Optional[str]:
    """Accept ISO datetime/date AND German dd.mm.yyyy. Returns yyyy-mm-dd."""
    s = (s or "").strip()
    if s in NULLISH:
        return None
    m = ISO_RE.match(s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = GER_RE.match(s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo}-{d}"
    return None  # e.g. the single month-only Amtsantritt "12.2021"


def classify_election_type(art: str) -> str:
    """'HaOB' -> Oberbürgermeisterwahl; 'Ha' / '-' / '' -> Bürgermeisterwahl.

    Case-insensitive: the file contains one lowercase 'ha' (Seeland 2016).
    Both hauptamtliche and ehrenamtliche Bürgermeister are Bürgermeisterwahlen,
    consistent with every other GERDA state.
    """
    a = (art or "").replace(" ", " ").strip().lower()
    return "Oberbürgermeisterwahl" if a == "haob" else "Bürgermeisterwahl"


def ratio(num, den) -> Optional[float]:
    if num is None or den in (None, 0):
        return None
    return num / den


def norm_name(s: str) -> str:
    """Fold a Gemeinde name for crosswalk lookup (umlauts, suffixes, punctuation)."""
    s = unicodedata.normalize("NFKD", (s or "").lower())
    s = s.replace("ß", "ss").replace("ä", "a").replace("ö", "o").replace("ü", "u")
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r",?\s*(stadt|hansestadt|flecken|lutherstadt)$", "", s)
    return re.sub(r"[^a-z0-9]", "", s)


def repair_source_rows(data: list) -> tuple:
    """Fix two defects in the source's Gemeinde/AGS bookkeeping.

    (A) DUPLICATE RECORDS. A handful of elections are listed twice — once under
        the Gemeinde's name at the time and once under the name of the Gemeinde
        it later merged into (e.g. 2008-03-30 "Sandersdorf" and
        "Sandersdorf-Brehna, Stadt": same AGS, date, winner, turnout and
        byte-identical candidate blocks; they differ only in the name label and
        in whether an Amtsantritt is filled). Rows identical on the whole
        election payload — everything except the two name columns and the
        Amtsantritt — are collapsed, keeping the record that carries the
        Amtsantritt date.

    (B) SHARED AGS. For several 2008/2009 (and a few 1994/2001) elections the
        "AGS am Wahltag" column holds the POST-merger code, so a group of then-
        independent Gemeinden all carry one AGS. Example: ten separate 2008-02-24
        Bürgermeisterwahlen (Breitenfeld, Mieste, Jävenitz, …) are all stamped
        15081135, the code of Gardelegen — which they only merged into in 2010.
        Left alone these collapse downstream, because stage 01 reduces by
        (ags, election_date, round) and would keep exactly one of the ten.

        The true historical AGS is recovered from `ags_crosswalks.csv` by
        (normalised Gemeinde name, election year). A match is accepted ONLY if
        it is unique for that year AND the candidate AGS harmonises to the same
        municipality the source itself reports (its ags_21 equals either the
        source's "AGS aktuell" or the ags_21 of the code the source gave). That
        guard rejects same-name Gemeinden elsewhere in the state — e.g. a 1994
        "Dannefeld" that harmonises to Gardelegen, where the source's Dannefeld
        belongs to Salzwedel.

    Returns (repaired_rows, stats) where stats["shared_flag"] holds the ids of
    rows still sharing an AGS after the repair.
    """
    # ---- (A) duplicate records ------------------------------------------
    # Identity of an ELECTION: who ran, how many votes, and the aggregate
    # turnout — deliberately excluding the Gemeinde name, the Amtsantritt, and
    # the party/title/gender fields, because the duplicate pairs differ only in
    # how completely those optional fields are filled (e.g. the two 1994
    # "Schenkenhorst" records are identical except that one gives the
    # candidate's party as CDU and the other leaves it blank).
    identity_idx = ([C_AGS_WT, C_AGS_AK, C_ERGART, C_HW_DATE, C_SW_DATE,
                     C_W_NAME, C_W_VORNAME, C_ART]
                    + list(range(C_HW_WBER, C_ANZ + 1))
                    + [FIRST_B_COL + B_WIDTH * k + off
                       for k in range(N_CAND_SLOTS) for off in (1, 2, 5, 6)])

    def richness(r) -> int:
        return sum(1 for i in range(len(r)) if cell(r, i) not in NULLISH)

    best: dict = {}
    order: list = []
    for r in data:
        k = tuple(cell(r, i) for i in identity_idx)
        if k not in best:
            best[k] = r
            order.append(k)
        elif richness(r) > richness(best[k]):
            best[k] = r          # keep whichever record carries more detail
    rows = [best[k] for k in order]
    n_dup = len(data) - len(rows)

    # ---- (B) shared AGS --------------------------------------------------
    groups: dict = {}
    for r in rows:
        groups.setdefault((cell(r, C_AGS_WT), cell(r, C_HW_DATE)[:10]), []).append(r)
    shared = [g for g in groups.values() if len(g) > 1]

    n_recovered = 0
    shared_flag: set = set()
    if shared:
        by_name: dict = {}
        to_21: dict = {}
        cw_path = os.path.join(ROOT, "data", "crosswalks", "final", "ags_crosswalks.csv")
        if os.path.exists(cw_path):
            with open(cw_path, encoding="utf-8", newline="") as fh:
                for row in csv.DictReader(fh):
                    a = row["ags"].zfill(8)
                    by_name.setdefault((norm_name(row["ags_name"]), row["year"]),
                                       set()).add(a)
                    to_21[(a, row["year"])] = row["ags_21"].zfill(8)
        else:
            sys.stderr.write(f"  ! crosswalk not found ({cw_path}); cannot repair "
                             f"shared AGS\n")

        for g in shared:
            for r in g:
                given = cell(r, C_AGS_WT)
                cur = cell(r, C_AGS_AK)
                year = cell(r, C_HW_DATE)[:4]
                hits = by_name.get((norm_name(cell(r, C_NAME_WT)), year), set())
                if len(hits) == 1:
                    cand = next(iter(hits))
                    if cand == given:
                        continue                       # already correct
                    t = to_21.get((cand, year), "")
                    if t and (t == cur or t == to_21.get((given, year), "")):
                        r[C_AGS_WT] = cand             # accept the recovery
                        n_recovered += 1
                        continue
                shared_flag.add(id(r))

        # anything still sharing an (ags, date) after the repair stays flagged
        regrouped: dict = {}
        for r in rows:
            regrouped.setdefault((cell(r, C_AGS_WT), cell(r, C_HW_DATE)[:10]),
                                 []).append(r)
        for g in regrouped.values():
            if len(g) > 1:
                for r in g:
                    shared_flag.add(id(r))
                    sys.stderr.write(
                        f"  ~ shared AGS retained: {cell(r, C_AGS_WT)} "
                        f"{cell(r, C_NAME_WT)} {cell(r, C_HW_DATE)[:10]}\n")
            else:
                shared_flag.discard(id(g[0]))

    return rows, {"n_dup": n_dup, "n_recovered": n_recovered,
                  "n_unresolved": len(shared_flag), "shared_flag": shared_flag}


def main() -> None:
    if not os.path.exists(IN_XLSX):
        sys.exit(f"missing input: {IN_XLSX}")

    wb = openpyxl.load_workbook(IN_XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"sheet {SHEET!r} not found; have {wb.sheetnames}")
    ws = wb[SHEET]
    raw = [[("" if c is None else str(c).strip()) for c in r]
           for r in ws.iter_rows(values_only=True)]

    # --- verify the header shape rather than trusting fixed offsets blindly ---
    r3, r4 = raw[2], raw[3]
    if cell(r4, C_AGS_WT).replace("\n", " ") != "AGS* am Wahltag":
        sys.exit(f"unexpected header at col {C_AGS_WT}: {cell(r4, C_AGS_WT)!r}")
    if cell(r4, C_AGS_AK).replace("\n", " ") != "AGS aktuell":
        sys.exit(f"unexpected header at col {C_AGS_AK}: {cell(r4, C_AGS_AK)!r}")
    bcols = [i for i, v in enumerate(r3) if re.fullmatch(r"B\d+", v)]
    if not bcols or bcols[0] != FIRST_B_COL:
        sys.exit(f"candidate blocks start at {bcols[:1]}, expected {FIRST_B_COL}")
    if len(bcols) > 1 and (bcols[1] - bcols[0]) != B_WIDTH:
        sys.exit(f"candidate block width {bcols[1] - bcols[0]}, expected {B_WIDTH}")
    if len(bcols) != N_CAND_SLOTS:
        sys.exit(f"found {len(bcols)} candidate blocks, expected {N_CAND_SLOTS}")

    # Data rows: col 0 must be an 8-digit AGS (drops the trailing footnote row).
    data = [r for r in raw[5:] if r and re.fullmatch(r"\d{8}", cell(r, 0))]
    if not data:
        sys.exit("no data rows found")

    data, fix = repair_source_rows(data)
    print(f"  source repair: {fix['n_dup']} duplicate record(s) dropped, "
          f"{fix['n_recovered']} shared-AGS rows re-coded to their true historical "
          f"AGS, {fix['n_unresolved']} left sharing an AGS (flagged)")
    shared_flag = fix["shared_flag"]

    out_rows: list[dict] = []
    n_hw = n_sw = n_sw_winner_only = n_slot_dup = n_sw_sameday = n_sw_spurious = 0
    n_by_office: dict[str, int] = {}
    warn_sum_hw = warn_sum_sw = 0
    warn_winner = 0
    n_anz_mismatch = 0

    for ridx, r in enumerate(data, start=6):
        ags = cell(r, C_AGS_WT)
        ags_cur = cell(r, C_AGS_AK)
        gname = cell(r, C_NAME_WT)
        gname_cur = cell(r, C_NAME_AK)

        if not ags.startswith(STATE):
            sys.exit(f"row {ridx}: non-ST AGS am Wahltag {ags!r}")
        if not re.fullmatch(r"15\d{6}", ags_cur):
            sys.exit(f"row {ridx}: bad AGS aktuell {ags_cur!r}")

        hw_date = parse_date(cell(r, C_HW_DATE))
        if hw_date is None:
            sys.exit(f"row {ridx} ({ags} {gname}): unparseable Hauptwahl date "
                     f"{cell(r, C_HW_DATE)!r}")
        sw_date = parse_date(cell(r, C_SW_DATE))
        if sw_date is not None and sw_date == hw_date:
            # 4 rows carry a Stichwahl date identical to the Hauptwahl date,
            # which cannot happen. Where votes exist they confirm no runoff was
            # needed (Ebendorf 1994: 257/503 = 51%; Eichenbarleben and Eickendorf:
            # single candidate at 100%). Treat the cycle as Hauptwahl-only so the
            # winner is seated on the Hauptwahl and no same-date duplicate round
            # is emitted.
            sys.stderr.write(f"  ~ row {ridx} {ags} {gname}: Stichwahl date equals "
                             f"Hauptwahl date ({hw_date}) — treated as no runoff\n")
            sw_date = None
            n_sw_sameday += 1

        election_type = classify_election_type(cell(r, C_ART))
        n_by_office[election_type] = n_by_office.get(election_type, 0) + 1

        wber_hw, waehl_hw = clean_num(cell(r, C_HW_WBER)), clean_num(cell(r, C_HW_WAEHL))
        ung_hw, guelt_hw = clean_num(cell(r, C_HW_UNG)), clean_num(cell(r, C_HW_GUELT))
        wber_sw, waehl_sw = clean_num(cell(r, C_SW_WBER)), clean_num(cell(r, C_SW_WAEHL))
        ung_sw, guelt_sw = clean_num(cell(r, C_SW_UNG)), clean_num(cell(r, C_SW_GUELT))

        anz = clean_num(cell(r, C_ANZ))
        anz = int(anz) if anz is not None else None

        w_last = clean_text(cell(r, C_W_NAME))
        w_first = clean_text(cell(r, C_W_VORNAME))
        w_gebjahr = clean_num(cell(r, C_W_GEBJAHR))
        amtsantritt = parse_date(cell(r, C_AMTSANTRITT))
        ergart = clean_text(cell(r, C_ERGART))

        # --- candidate slots ---------------------------------------------
        cands = []
        for k in range(N_CAND_SLOTS):
            b = FIRST_B_COL + B_WIDTH * k
            titel = clean_text(cell(r, b + 0))
            last = clean_text(cell(r, b + 1))
            first = clean_text(cell(r, b + 2))
            partei = clean_text(cell(r, b + 3))
            geschl = clean_text(cell(r, b + 4))
            v_hw = clean_num(cell(r, b + 5))
            v_sw = clean_num(cell(r, b + 6))
            # A slot is real if it names someone, is an explicitly-unknown
            # candidate ('.'), or carries votes. '-' everywhere = empty slot.
            raw_name = cell(r, b + 1).strip()
            occupied = bool(last or partei or titel or geschl) or raw_name == "." \
                or v_hw is not None or v_sw is not None
            if not occupied:
                continue
            cands.append(dict(titel=titel, last=last, first=first, partei=partei,
                              geschl=geschl, v_hw=v_hw, v_sw=v_sw))

        # One row (Zörnigall 1994) lists the same person in two candidate slots.
        # Nobody stands twice in one election, so collapse identically-named
        # slots and keep whichever carries the vote detail.
        def vote_detail(x) -> int:
            return (x["v_hw"] is not None) + (x["v_sw"] is not None)

        collapsed: list = []
        for c in cands:
            prior = next((p for p in collapsed
                          if c["last"]
                          and p["last"].casefold() == c["last"].casefold()
                          and p["first"].casefold() == c["first"].casefold()), None)
            if prior is None:
                collapsed.append(c)
                continue
            n_slot_dup += 1
            sys.stderr.write(f"  ~ row {ridx} {ags} {gname} {hw_date}: candidate "
                             f"{c['last']!r} appears in two slots — collapsed\n")
            if vote_detail(c) > vote_detail(prior):
                collapsed[collapsed.index(prior)] = c
        cands = collapsed

        if anz is not None and anz != len(cands):
            n_anz_mismatch += 1
            sys.stderr.write(
                f"  ! row {ridx} {ags} {gname} {hw_date}: Anzahl Wahlvorschläge="
                f"{anz} but {len(cands)} slots populated\n")

        # --- vote integrity ------------------------------------------------
        sum_hw = sum(c["v_hw"] for c in cands if c["v_hw"] is not None)
        if guelt_hw is not None and sum_hw > 0 and abs(sum_hw - guelt_hw) > 0.5 \
                and all(c["v_hw"] is not None for c in cands):
            warn_sum_hw += 1
            sys.stderr.write(f"  ! row {ridx} {ags} {gname} {hw_date}: HW sum "
                             f"{sum_hw:.0f} != Gültige {guelt_hw:.0f}\n")
        sum_sw = sum(c["v_sw"] for c in cands if c["v_sw"] is not None)
        if guelt_sw is not None and sum_sw > 0 and abs(sum_sw - guelt_sw) > 0.5:
            warn_sw_ok = all(c["v_sw"] is not None
                             for c in cands if c["v_sw"] is not None)
            if warn_sw_ok:
                warn_sum_sw += 1
                sys.stderr.write(f"  ! row {ridx} {ags} {gname} {sw_date}: SW sum "
                                 f"{sum_sw:.0f} != Gültige {guelt_sw:.0f}\n")

        # A Stichwahl date with NO candidate-level Stichwahl votes, on a cycle
        # whose Hauptwahl already produced an absolute majority, is
        # contradictory — the Hauptwahl seated the mayor, so the stray runoff
        # date is not usable. 8 rows, e.g. Klostermansfeld 2013 where the sole
        # candidate took 604/604 = 100% on 13.01. yet a runoff is dated 03.02.,
        # and Calbe (Saale) 1994 where Schacke won 75.1% outright. Treat these
        # cycles as Hauptwahl-only so the winner is seated on the Hauptwahl.
        if sw_date is not None and guelt_hw and \
                not any(c["v_sw"] is not None for c in cands):
            top_hw = max((c["v_hw"] for c in cands if c["v_hw"] is not None),
                         default=None)
            if top_hw is not None and top_hw / guelt_hw > 0.5:
                sys.stderr.write(
                    f"  ~ row {ridx} {ags} {gname} {hw_date}: Stichwahl dated "
                    f"{sw_date} but the Hauptwahl was won outright "
                    f"({top_hw / guelt_hw:.1%}) and carries no runoff votes "
                    f"— treated as no runoff\n")
                sw_date = None
                n_sw_spurious += 1

        # --- winner identification -----------------------------------------
        # Prefer StaLA's explicit row-level "Wahlsieger" name (authoritative even
        # when a Ja/Nein round polls <50%); fall back to top votes on the
        # decisive round when the winner is un-named (common in 1994).
        decisive_key = "v_sw" if sw_date is not None else "v_hw"
        is_win = [False] * len(cands)
        if w_last:
            hits = [i for i, c in enumerate(cands)
                    if c["last"].casefold() == w_last.casefold()
                    and (not w_first or not c["first"]
                         or c["first"].casefold() == w_first.casefold())]
            if len(hits) == 1:
                is_win[hits[0]] = True
            elif len(hits) > 1:
                # disambiguate by top votes among the same-name hits
                best = max(hits, key=lambda i: (cands[i][decisive_key] or -1))
                is_win[best] = True
                warn_winner += 1
                sys.stderr.write(f"  ! row {ridx} {ags} {gname} {hw_date}: winner "
                                 f"{w_last!r} matched {len(hits)} slots\n")
        if not any(is_win) and cands:
            scored = [i for i, c in enumerate(cands) if c[decisive_key] is not None]
            if scored:
                is_win[max(scored, key=lambda i: cands[i][decisive_key])] = True

        common = dict(
            ags=ags, ags_name=gname, state=STATE, state_name=STATE_NAME,
            election_type=election_type, source_url=SOURCE_TAG,
            ags_current=ags_cur, ags_name_current=gname_cur,
            ergebnisart=ergart, amtsantritt_date=amtsantritt or "",
            winner_birth_year=int(w_gebjahr) if w_gebjahr is not None else "",
            # TRUE where the source still gives several distinct Gemeinden the
            # same AGS for this date, so stage 01 will reduce them to one.
            flag_shared_ags=id(r) in shared_flag,
        )

        # --- Hauptwahl rows --------------------------------------------------
        turnout_hw = ratio(waehl_hw, wber_hw)
        for c, win in zip(cands, is_win):
            out_rows.append(dict(
                common,
                election_year=int(hw_date[:4]), election_date=hw_date,
                round="hauptwahl",
                eligible_voters=wber_hw, number_voters=waehl_hw,
                valid_votes=guelt_hw, invalid_votes=ung_hw, turnout=turnout_hw,
                candidate_name=f"{c['last']}, {c['first']}".strip(", "),
                candidate_last_name=c["last"], candidate_first_name=c["first"],
                candidate_title=c["titel"], candidate_party=c["partei"],
                candidate_gender=c["geschl"],
                candidate_votes=c["v_hw"],
                candidate_voteshare=ratio(c["v_hw"], guelt_hw),
                # decisive round only, mirroring the previous parser
                is_winner=bool(win and sw_date is None),
            ))
        n_hw += 1

        # --- Stichwahl rows ---------------------------------------------------
        if sw_date is not None:
            turnout_sw = ratio(waehl_sw, wber_sw)

            def sw_row(c, win, votes):
                return dict(
                    common,
                    election_year=int(sw_date[:4]), election_date=sw_date,
                    round="stichwahl",
                    eligible_voters=wber_sw, number_voters=waehl_sw,
                    valid_votes=guelt_sw, invalid_votes=ung_sw, turnout=turnout_sw,
                    candidate_name=f"{c['last']}, {c['first']}".strip(", "),
                    candidate_last_name=c["last"], candidate_first_name=c["first"],
                    candidate_title=c["titel"], candidate_party=c["partei"],
                    candidate_gender=c["geschl"],
                    candidate_votes=votes,
                    candidate_voteshare=ratio(votes, guelt_sw),
                    is_winner=bool(win),
                )

            advanced = [(c, w) for c, w in zip(cands, is_win) if c["v_sw"] is not None]
            if advanced:
                for c, win in advanced:
                    out_rows.append(sw_row(c, win, c["v_sw"]))
                n_sw += 1
            else:
                # A runoff demonstrably took place (the sheet gives its date and
                # names the Wahlsieger) but no candidate-level Stichwahl votes
                # were reported — 75 of these are 1994, which the Erläuterung
                # flags as winner-only. Record the round and its winner rather
                # than silently dropping the runoff; votes stay NA.
                widx = next((i for i, w in enumerate(is_win) if w), None)
                if widx is not None:
                    out_rows.append(sw_row(cands[widx], True, None))
                    n_sw_winner_only += 1

    if not out_rows:
        sys.exit("no rows emitted — abort")

    # =====================================================================
    # MERGE with the bmbm.csv extract (st_bmbm_parsed.csv).
    #
    # Priority is bmbm-first for the elections bmbm covers, because bmbm is
    # demonstrably the more accurate record per row, while the historical file
    # is (per its own Erläuterung) "in der Aufbauphase". Evidence found by
    # diffing the two over their 270 shared election-rounds:
    #   * ~16 surname typos in the historical file that bmbm has right (single
    #     dropped or transposed letters — the same person, same votes, spelled
    #     two ways).
    #   * Arnstein 2023: the historical file swaps the surname and forename
    #     fields for one candidate; bmbm has them the right way round.
    #   * Halle 2025-02-02: the two sources differ by 30 on one candidate's vote
    #     count — only bmbm's figure makes the round add up
    #     (87407 + 30 = 87437 = Gültige Stimmen).
    #   * Landsberg 2022 runoff: historical lists Schenk (3rd in the HW with
    #     1047) as the runoff participant; the top two were Müller (1533) and
    #     Halfpap (1275). bmbm correctly has Müller.
    #   * Oebisfelde-Weferlingen 2023: historical has NO runoff and seats
    #     Blanck off a 1762/5488 = 32% Hauptwahl; bmbm has the real 2023-10-08
    #     runoff that Blanck won.
    # Conversely the historical file fixes bmbm's one known corruption (the
    # Genthin 2024 B08/B09 record → its 8th candidate, 96 votes), so historical
    # candidates that bmbm lacks are grafted back in.
    #
    # Three rules:
    #  (1) bmbm rows are taken verbatim for every (ags, election_date) it covers.
    #  (2) A historical election-round at a date bmbm does NOT cover is DROPPED
    #      as a date typo iff it is byte-identical in shape to a bmbm round for
    #      the same Gemeinde at a different date — same round, same Gültige, and
    #      the same multiset of candidate votes. That fingerprint is strict
    #      enough to reject coincidences (e.g. Allstedt 2009 vs 2024 share
    #      Gültige=3434 but no vote multiset, so they are correctly kept apart).
    #      This catches 4 real typos, incl. Osternienburger Land's runoff dated
    #      2024-09-24 whose Hauptwahl was 2023-09-10.
    #  (3) Everything else from the historical file is kept — that is the whole
    #      1994-2026 back-catalogue bmbm never had.
    # =====================================================================
    hist_rows = out_rows
    cols = list(hist_rows[0].keys())

    def norm_num(x) -> str:
        """Comparable string for a numeric field across CSV/float sources."""
        if x is None or x == "":
            return ""
        try:
            return str(int(float(x)))
        except (TypeError, ValueError):
            return str(x).strip()

    bmbm_rows: list[dict] = []
    if os.path.exists(IN_BMBM):
        with open(IN_BMBM, encoding="utf-8", newline="") as fh:
            for row in csv.DictReader(fh):
                out = {k: row.get(k, "") for k in cols}
                out["source_url"] = BMBM_TAG
                bmbm_rows.append(out)
    else:
        sys.stderr.write(f"  ! bmbm source not found: {IN_BMBM} "
                         f"(run 00_st_stala_parse.py first)\n")

    def round_sig(rows) -> tuple:
        r0 = rows[0]
        return (r0["round"], norm_num(r0["valid_votes"]),
                tuple(sorted(norm_num(c["candidate_votes"]) for c in rows)))

    def claim_keys(row) -> set:
        """Identity claims used for candidate-level dedup (ANY match = dup)."""
        out = set()
        for f in ("candidate_last_name", "candidate_first_name"):
            v = str(row.get(f) or "").strip().casefold()
            if v:
                out.add(("N", v))
        v = norm_num(row.get("candidate_votes"))
        if v:
            out.add(("V", v))
        return out

    bmbm_pairs = {(r["ags"], r["election_date"]) for r in bmbm_rows}
    bmbm_by_round: dict = {}
    for r in bmbm_rows:
        bmbm_by_round.setdefault((r["ags"], r["election_date"], r["round"]), []).append(r)
    bmbm_sigs: dict = {}
    for k, rows in bmbm_by_round.items():
        bmbm_sigs.setdefault(k[0], []).append((round_sig(rows), k[1]))

    hist_by_round: dict = {}
    for r in hist_rows:
        hist_by_round.setdefault((r["ags"], r["election_date"], r["round"]), []).append(r)

    out_rows = list(bmbm_rows)
    n_graft = n_phantom = n_hist_rounds = 0
    for k in sorted(hist_by_round):
        ags, date, rnd = k
        rows = hist_by_round[k]
        if (ags, date) in bmbm_pairs:
            # (1) shared election — bmbm wins; graft only candidates it lacks.
            claims: set = set()
            for p in bmbm_by_round.get(k, []):
                claims |= claim_keys(p)
            for r in rows:
                if claim_keys(r) & claims:
                    continue
                out_rows.append(r)
                n_graft += 1
                sys.stderr.write(
                    f"  + grafted candidate missing from bmbm: {ags} "
                    f"{r['ags_name']} {date} {rnd} "
                    f"{r['candidate_last_name']!r} votes={norm_num(r['candidate_votes'])}\n")
            continue
        # (2) phantom date-typo?
        sig = round_sig(rows)
        dup = [d for (s, d) in bmbm_sigs.get(ags, ()) if s == sig]
        if dup:
            n_phantom += 1
            sys.stderr.write(
                f"  - dropped date-typo round: {ags} {rows[0]['ags_name']} "
                f"{date} {rnd} duplicates bmbm's {dup[0]}\n")
            continue
        # (3) genuine historical election
        out_rows.extend(rows)
        n_hist_rounds += 1
    with open(OUT_CSV, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for row in out_rows:
            for k in ("eligible_voters", "number_voters", "valid_votes",
                      "invalid_votes", "candidate_votes"):
                v = row.get(k)
                if isinstance(v, float) and v == int(v):
                    row[k] = int(v)
            w.writerow(row)

    years = sorted({str(r["election_date"])[:4] for r in out_rows})
    n_elections = len({(r["ags"], r["election_date"], r["round"]) for r in out_rows})
    print(f"wrote {OUT_CSV}")
    print(f"  historical file: {n_hw} Hauptwahl + {n_sw} Stichwahl "
          f"(+{n_sw_winner_only} runoffs recorded winner-only, no vote detail)")
    print(f"  merge: {len(bmbm_rows)} bmbm rows kept verbatim (authoritative for "
          f"their elections), {n_hist_rounds} historical rounds added, "
          f"{n_graft} candidates grafted into bmbm elections, "
          f"{n_phantom} date-typo rounds dropped")
    print(f"  {len(out_rows)} candidate rows across {n_elections} "
          f"election-rounds ({years[0]}-{years[-1]})")
    print(f"  distinct Gemeinden: AGS am Wahltag="
          f"{len({r['ags'] for r in out_rows})}, "
          f"AGS aktuell={len({r['ags_current'] for r in out_rows if r['ags_current']})}")
    for et, n in sorted(n_by_office.items()):
        print(f"  {et}: {n} elections (historical file)")
    if warn_sum_hw or warn_sum_sw:
        print(f"  vote-integrity mismatches (SOURCE anomalies): "
              f"{warn_sum_hw} HW, {warn_sum_sw} SW")
    else:
        print("  vote-integrity: OK")
    if n_anz_mismatch:
        print(f"  Anzahl-Wahlvorschläge mismatches: {n_anz_mismatch} "
              f"(source-flagged: the Erläuterung warns these may be inconsistent)")


if __name__ == "__main__":
    main()
