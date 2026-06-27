#!/usr/bin/env python3
"""Stage-0 parser for the Bayern Kommunalwahl of 8 March 2026 (Stichwahl 22 March).

Source: Bayerisches Landesamt für Statistik, "Liste der Ersten Bürgermeister/-innen,
der Oberbürgermeister/-innen sowie der Landrätinnen und Landräte mit Wahlergebnissen"
(`wahlergebnisse_mandatsräger.xlsx`, Stand 24.06.2026) saved at
data/mayoral_elections/raw/bayern/BY_Kommunalwahl2026_Mandatstraeger_stand20260624.xlsx

The file is a current-officeholder snapshot (the most recent election that seated
each sitting Bürgermeister/OB/Landrat in Bayern, Wahltage ~2020-2026). GERDA's main
Bayern source ("Wahlen seit 1945", Stand 2025-11) already covers everything through
2025, so THIS parser keeps only the new **2026** elections (the 8/22 March 2026
Kommunalwahl plus any 2026 by-elections) to avoid duplicating existing rows.

Two sheets are joined by AGS:
  * "Wahlergebnisse"  — candidate-level: per (AGS, Wahltag, [Stichwahl]) the
    Stimmberechtigte / Wähler / Wahlbeteiligung / ungültige / gültige Stimmen and
    up to 13 Wahlvorschläge (Kennwort = party/coalition + gültige Stimmen = votes).
  * "Mandatsträger"   — the elected person per AGS: Amtsbezeichnung, Name,
    Geschlecht, Geburtsjahr, Erster Amtsantritt, Stimmenanteil, Wahlvorschlag.

Office classifier (existing Bayern rule = Amtsbezeichnung / Gebietsart):
  Amtsbezeichnung contains "Landrat"          -> Landratswahl (split to landrat ds)
  Amtsbezeichnung contains "Oberbürgermeister" -> Oberbürgermeisterwahl
  otherwise (Erster Bürgermeister)            -> Bürgermeisterwahl

Output (candidate-level long; MV/BB/ST/HE-compatible, + Bayern winner name/gender/
birth year/Amtsantritt attached to the winning Wahlvorschlag):
  data/mayoral_elections/raw/bayern/by2026_parsed.csv
Run:  python3 code/mayoral_elections/00_by_kommunalwahl2026_parse.py
"""

import csv
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
RAW_DIR = os.path.join(ROOT, "data", "mayoral_elections", "raw", "bayern")
XLSX = os.path.join(RAW_DIR, "BY_Kommunalwahl2026_Mandatstraeger_stand20260624.xlsx")
OUT = os.path.join(RAW_DIR, "by2026_parsed.csv")

STATE, STATE_NAME = "09", "Bayern"
YEAR = "2026"   # keep only elections held in 2026 (new vs the Wahlen-seit-1945 file)


def iso(dt):
    if dt is None:
        return ""
    s = str(dt)
    return s[:10]   # 'YYYY-MM-DD 00:00:00' -> 'YYYY-MM-DD'


def to_int(x):
    if x is None:
        return None
    try:
        return int(float(str(x).replace(",", ".")))
    except ValueError:
        return None


def to_floatshare(x):
    if x is None:
        return None
    try:
        return float(str(x).replace(",", ".")) / 100.0
    except ValueError:
        return None


def clean_party(p):
    """Normalise the Kennwort. Empty markers -> '' (parteilos/Einzelbewerber)."""
    p = " ".join(str(p or "").split())
    return "" if p in ("--", "–", "—", ".", "") else p


def classify(amtsbez, gebietsart):
    a = (amtsbez or "")
    if "Landrat" in a or (gebietsart or "") == "Landkreis":
        return "Landratswahl"
    if "Oberbürgermeister" in a or (gebietsart or "") == "kreisfreie Stadt":
        return "Oberbürgermeisterwahl"
    return "Bürgermeisterwahl"


FIELDS = ["ags", "ags_name", "state", "state_name", "election_year", "election_date",
          "election_type", "round", "eligible_voters", "number_voters", "valid_votes",
          "invalid_votes", "turnout", "candidate_name", "candidate_party",
          "candidate_votes", "candidate_voteshare", "candidate_gender",
          "candidate_birth_year", "is_winner", "candidate_rank", "n_candidates",
          "amtsantritt", "source_file"]


def main():
    if not os.path.exists(XLSX):
        raise SystemExit(f"Raw XLSX not found: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # ---- sheet 1: Mandatsträger (winner per AGS, only 2026) ----
    winners = {}
    ws1 = wb["Mandatsträger"]
    for r in ws1.iter_rows(min_row=3, values_only=True):
        if not r[0] or not r[5] or not iso(r[5]).startswith(YEAR):
            continue
        ags = STATE + str(r[0]).strip()
        winners[ags] = {
            "amtsbez": str(r[7] or "").strip(),
            "name": str(r[8] or "").strip(),
            "gender": {"m": "male", "w": "female", "f": "female"}.get(str(r[9] or "").strip().lower(), ""),
            "birthyear": to_int(r[10]),
            "amtsantritt": iso(r[6]),
        }

    # ---- sheet 2: Wahlergebnisse (candidate-level, only 2026) ----
    ws2 = wb["Wahlergebnisse"]
    rows_by_ags = {}   # ags -> list of (round-dict)
    for r in ws2.iter_rows(min_row=3, values_only=True):
        if not r[0] or not r[5] or not iso(r[5]).startswith(YEAR):
            continue
        ags = STATE + str(r[0]).strip()
        gebietsart = str(r[2] or "").strip()
        valid = to_int(r[11])
        cands = []
        for j in range(12, len(r) - 1, 2):       # (Kennwort, gültige Stimmen) pairs
            v = to_int(r[j + 1])
            if v is None:                         # no votes -> trailing empty slot
                continue
            cands.append({
                "party": clean_party(r[j]),
                "votes": v,
                "share": (v / valid) if valid else None,
            })
        rows_by_ags.setdefault(ags, []).append({
            "ags_name": str(r[1] or "").strip(),
            "gebietsart": gebietsart,
            "date": iso(r[5]),
            "round": "stichwahl" if (str(r[6] or "").strip() == "Stichwahl") else "hauptwahl",
            "eligible": to_int(r[7]),
            "voters": to_int(r[8]),
            "turnout": to_floatshare(r[9]),
            "invalid": to_int(r[10]),
            "valid": valid,
            "cands": cands,
        })

    out = []
    n_elections = n_ob = n_lr = 0
    for ags, rounds in rows_by_ags.items():
        win = winners.get(ags, {})
        etype = classify(win.get("amtsbez"), rounds[0]["gebietsart"])
        n_elections += 1
        n_ob += (etype == "Oberbürgermeisterwahl")
        n_lr += (etype == "Landratswahl")
        # decisive round = the chronologically last one (Stichwahl if held)
        decisive = max(rounds, key=lambda x: x["date"])
        dec_ranked = sorted(decisive["cands"], key=lambda c: -c["votes"])
        win_kennwort = dec_ranked[0]["party"] if dec_ranked else ""
        for rec in rounds:
            is_dec = rec is decisive
            # rank the Wahlvorschläge by votes; winner = rank 1 in the decisive round
            # (use object identity so equal-vote / empty-party rows aren't conflated)
            ranked = sorted(rec["cands"], key=lambda c: -c["votes"])
            rank_by_id = {id(c): i + 1 for i, c in enumerate(ranked)}
            ncand = len(rec["cands"])
            for c in rec["cands"]:
                rank = rank_by_id[id(c)]
                is_winner = is_dec and rank == 1
                # Attach the elected person's name/gender/birth year to THEIR
                # Wahlvorschlag in every round (matched by Kennwort) so the Hauptwahl
                # and Stichwahl rows pair by name in the wide reshaping. For an empty
                # Kennwort (independent) we can't match across rounds -> name on the
                # decisive winner only.
                is_person = (c["party"] == win_kennwort and win_kennwort != "") or \
                            (win_kennwort == "" and is_winner)
                out.append({
                    "ags": ags, "ags_name": rec["ags_name"], "state": STATE,
                    "state_name": STATE_NAME, "election_year": rec["date"][:4],
                    "election_date": rec["date"], "election_type": etype,
                    "round": rec["round"],
                    "eligible_voters": rec["eligible"] if rec["eligible"] is not None else "",
                    "number_voters": rec["voters"] if rec["voters"] is not None else "",
                    "valid_votes": rec["valid"] if rec["valid"] is not None else "",
                    "invalid_votes": rec["invalid"] if rec["invalid"] is not None else "",
                    "turnout": rec["turnout"] if rec["turnout"] is not None else "",
                    # Bayern publishes no candidate names; the Mandatsträger sheet
                    # gives the elected person, attached to the winning Wahlvorschlag.
                    "candidate_name": win.get("name", "") if is_person else "",
                    "candidate_party": c["party"],
                    "candidate_votes": c["votes"] if c["votes"] is not None else "",
                    "candidate_voteshare": c["share"] if c["share"] is not None else "",
                    "candidate_gender": win.get("gender", "") if is_person else "",
                    "candidate_birth_year": win.get("birthyear", "") if is_person else "",
                    "is_winner": "TRUE" if is_winner else "FALSE",
                    "candidate_rank": rank,
                    "n_candidates": ncand,
                    "amtsantritt": win.get("amtsantritt", "") if is_person else "",
                    "source_file": os.path.basename(XLSX),
                })

    sys.stderr.write(
        f"\n=== Bayern Kommunalwahl 2026 parse ===\n"
        f"  2026 elections:   {n_elections}\n"
        f"    Oberbürgermeister: {n_ob}\n"
        f"    Landrat:           {n_lr}\n"
        f"    Bürgermeister:     {n_elections - n_ob - n_lr}\n"
        f"  candidate rows:   {len(out)}\n"
        f"  Mandatsträger 2026: {len(winners)}\n")
    no_winner = [a for a in rows_by_ags if a not in winners]
    if no_winner:
        sys.stderr.write(f"  NOTE: {len(no_winner)} elections without a Mandatsträger match\n")

    with open(OUT, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(out)
    sys.stderr.write(f"  Wrote {len(out)} rows -> {OUT}\n")


if __name__ == "__main__":
    main()
