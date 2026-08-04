#!/usr/bin/env python3
"""
Stage 0 — Brandenburg hauptamtliche Bürgermeisterwahlen, 2010-2026.

PRIMARY SOURCE
    data/mayoral_elections/raw/brandenburg/
        BB_Buergermeisterwahlen_ab2010_Landeswahlleiter.xlsx

    Supplied on request by the Geschäftsstelle des Landeswahlleiters,
    Ministerium des Innern und für Kommunales des Landes Brandenburg
    (Christopher Sokol, 31.07.2026; sheet "Inhalte_Ergebnisse_20260731_093").
    One WIDE row per (Gemeinde, Wahltag, Wahlart): 8 metadata columns, the two
    runoff-qualifier labels, then 10 candidate blocks of 5 columns
    (bew_titel / name_bewerber / art / name_wahlvorschlagstraeger /
    stimmenanzahl), then 5 "sieger" columns naming the elected person.

    The cover letter states two limits:
      * results are on the CURRENT Gebietsstand (01.01.2026) — historical
        boundaries are NOT reproduced, so BB rows (like Rheinland-Pfalz council
        rows) are already harmonised at source rather than on election-year
        boundaries;
      * digital transmission by the Gemeinden only became compulsory in 2014,
        so 2010-2013 makes no claim to completeness.
    No separate Briefwahl results exist.

FALLBACK SOURCE
    bb_bm_parsed.csv — the Landeswahlleiter web-portal scrape (00_bb_scrape.py),
    which covers only the current cycle. It is used for the (ags, date, round)
    keys the XLSX does not carry (currently 2: Heideblick and Groß Kreutz
    (Havel), both 2019-09-01).

OUTPUT
    data/mayoral_elections/raw/brandenburg/bb_lwl_parsed.csv
    Candidate-level long — one row per candidate per round — in the same schema
    the 01/01b Brandenburg blocks already consume, plus `is_winner`,
    `candidate_title`, `wahlvorschlag_art` and `source`.

SOURCE DEFECTS REPAIRED HERE (all reported, never silent)
  (a) One election is listed twice, byte-identical except that one copy leaves
      the `art_N` (Wahlvorschlagsart) columns empty — Schwielowsee 2018-09-30.
      Deduped on the result fingerprint, keeping the richer copy.
  (b) One election is listed twice under two different dates, one of which is
      not a Sunday — Tauche, 2021-03-21 (Sunday) and 2019-03-21 (Thursday),
      identical Wahlberechtigte/Wähler/gültige and identical candidate votes.
      The Landeswahlleiter's own portal publishes it as ~h_21032021_12067493,
      so the Thursday copy is dropped (same fingerprint rule as the
      Sachsen-Anhalt parser).
  (c) One row is a "Wahl durch Vertretung" (Falkenberg/Elster, 2020-01-23) —
      the Stadtverordnetenversammlung elected the mayor because no valid direct
      election happened. It carries no AGS, no counts and no name, so there is
      nothing to publish; it is dropped and reported.
  (d) Two 2026 Nordwestuckermark rows carry the Gemeinde name with a polling
      annotation appended ("Nordwestuckermark  OT Schönermark"); the " OT ..."
      suffix is stripped so the AGS has one name.
  (e) Nine candidates are spelled differently in the Hauptwahl and the Stichwahl
      of the same cycle. 01b pairs the two rounds BY NAME, so each of these
      split ONE person into TWO published rows — and in five cases the split row
      was the WINNER, who then carried the runoff result with no first-round
      votes while the first-round leader stood ranked as a loser. See
      reconcile_rounds() / NAME_FIXES below.

SOURCE ANOMALIES KEPT AS PUBLISHED
  * Seddiner See's 2022 Stichwahl is dated 22.02.2022, a Tuesday, in BOTH this
    file and the Landeswahlleiter portal (~s_22022022_12069596, whose result
    page prints "Wahltag 22.02.2022"). Two independent official records agree,
    and GERDA already publishes that date, so it is NOT silently moved to the
    20.02. that § 74 BbgKWahlG would imply.
  * Zehdenick was decided twice in 14 months (Stichwahl 2025-03-16 and a fresh
    Hauptwahl 2026-05-10). The file gives no annulment marker, so both stand.

Usage:  python3 code/mayoral_elections/00_bb_lwl_parse.py
"""

import csv
import datetime as dt
import difflib
import os
import re
import sys
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip3 install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
RAW_DIR = os.path.join(ROOT, "data", "mayoral_elections", "raw", "brandenburg")
XLSX = os.path.join(RAW_DIR, "BB_Buergermeisterwahlen_ab2010_Landeswahlleiter.xlsx")
PORTAL_CSV = os.path.join(RAW_DIR, "bb_bm_parsed.csv")
OUT_CSV = os.path.join(RAW_DIR, "bb_lwl_parsed.csv")

STATE, STATE_NAME = "12", "Brandenburg"

# Brandenburg has exactly four Oberbürgermeister, one per kreisfreie Stadt. The
# four Große kreisangehörige Städte (Eberswalde, Schwedt/Oder, Fürstenwalde,
# Königs Wusterhausen) do NOT: the Landeswahlleiter's own result pages head
# them "Ergebnis der Bürgermeisterwahl in ...". Same rule as 00_bb_scrape.py.
KREISFREIE = {"12051000", "12052000", "12053000", "12054000"}

N_CAND_BLOCKS = 10

# "Dr.", "Dr. phil.", "Dr.-Ing.", "Prof." … appear either before the surname
# ("Dr. Radant, Jana") or before the given name ("Tittel, Dr. Claudia").
TITLE_RE = re.compile(r"^((?:Dr|Prof)\.?(?:\s*-?\s*[A-Za-zÄÖÜäöü]+\.)*\.?)\s+")


def norm_ws(x):
    return re.sub(r"\s+", " ", str(x)).strip() if x is not None else ""


def strip_title(s):
    """Return (name_without_title, title_or_'')."""
    s = norm_ws(s)
    title = ""
    m = TITLE_RE.match(s)
    while m:
        title = (title + " " + m.group(1)).strip()
        s = s[m.end():].strip()
        m = TITLE_RE.match(s)
    return s, title


def split_name(raw):
    """'Tittel, Dr. Claudia' -> ('Tittel, Claudia', 'Tittel', 'Claudia', 'Dr.')"""
    s, t1 = strip_title(raw)
    if "," in s:
        last, first = [x.strip() for x in s.split(",", 1)]
    else:
        last, first = s, ""
    first, t2 = strip_title(first)
    last, t3 = strip_title(last)
    title = " ".join(x for x in (t1, t2, t3) if x).strip()
    full = f"{last}, {first}".strip().rstrip(",").strip()
    return full, last, first, title


def match_key(last, first):
    """Loose person key: surname + first token of the given name, casefolded.

    The `sieger_vorname` column drops second given names the candidate block
    keeps ("Fred" vs "Fred Hans Willi"), so a full-string compare would report
    six false winner mismatches.
    """
    return (last.casefold(), (first.split() or [""])[0].casefold())


def ars_to_ags(v):
    """12-digit ARS -> 8-digit AGS: Land+RB+Kreis (5) + Gemeinde (last 3).

    The 4 middle digits are the Gemeindeverbandsschlüssel and are NOT part of
    the AGS — for most amtsfreie Gemeinden they simply repeat the Gemeinde
    number, but Schwedt/Oder (5051) and Oberuckersee (5306) sit in a
    Verwaltungsgemeinschaft and do not.
    """
    s = str(v).strip()
    if not s.isdigit() or len(s) != 12:
        return None
    return s[:5] + s[9:12]


def load_xlsx():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = next(w for w in wb.worksheets if w.title.startswith("Inhalte_Ergebnisse"))
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    need = ["Wahlart", "AGS/ARS", "Gemeinde", "Datum der Wahl", "Wahlberechtigte",
            "Waehler", "ungueltige Stimmen", "gueltige Stimmen",
            "sieger_vorname", "sieger_nachname", "sieger_wahlvorschlagstraeger"]
    missing = [c for c in need if c not in hdr]
    if missing:
        sys.exit(f"FATAL: expected columns missing from {ws.title}: {missing}")
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v in (None, "") for v in vals):
            continue
        d = dict(zip(hdr, vals))
        d["_row"] = r
        d["_sheet"] = ws.title
        rows.append(d)
    return rows


def candidates_of(rec):
    out = []
    for i in range(1, N_CAND_BLOCKS + 1):
        raw = rec.get(f"name_bewerber_{i}")
        if raw in (None, ""):
            continue
        votes = rec.get(f"stimmenanzahl_{i}")
        out.append({
            "raw_name": norm_ws(raw),
            "party": norm_ws(rec.get(f"name_wahlvorschlagstraeger_{i}")),
            "art": norm_ws(rec.get(f"art_{i}")),
            "title_col": norm_ws(rec.get(f"bew_titel_{i}")),
            "votes": int(votes) if isinstance(votes, (int, float)) else None,
        })
    return out


def fingerprint(rec):
    """Identity of an election RESULT, independent of the date it is filed under."""
    cands = candidates_of(rec)
    return (norm_ws(rec["Wahlart"]), rec["Wahlberechtigte"], rec["Waehler"],
            rec["gueltige Stimmen"],
            tuple(sorted(c["votes"] for c in cands if c["votes"] is not None)))


# ---------------------------------------------------------------------------
# Cross-round name reconciliation
# ---------------------------------------------------------------------------
# The source spells the SAME person differently in the Hauptwahl and the
# Stichwahl of one cycle in 9 of 135 runoffs — a dropped letter (Kulike /
# Kulicke, Trebling / Tebling, Heinz-Georg / Hein-Georg, Herinrich / Heinrich,
# Maass / Maaß), a truncated surname ("K, Bengt" for "Kanzler, Bengt"), swapped
# name fields ("Ronny, Kretschmer"), or a shortened given name (Fred /
# Fred Hans Willi). 01b pairs the two rounds BY NAME, so each of these split one
# candidate into two rows — and in five cases the split row was the WINNER, who
# then carried the runoff result with no first-round votes while the first-round
# leader was left ranked as a loser.
#
# The reconciliation is anchored on the source's own runoff-qualifier columns
# (`kandidat0_stichwahl` / `kandidat1_stichwahl` on the Hauptwahl row), which
# name exactly who reached the runoff, in the Hauptwahl's spelling. Assignment
# is by exact match first, then similarity, then elimination — never by
# similarity alone, so two different people are never merged.
QUALIFIER_RE = re.compile(r"^\s*\d*\s*\.?\s*(.+?)\s*\((.*)\)\s*$")

# Where the two rounds disagree, which spelling is right is a question about a
# real person, not something a rule can decide: the Hauptwahl is right in some
# of these and the Stichwahl in others. Each entry below was checked against a
# source outside the election file (municipal websites, local press, the
# officeholder's own pages) and carries that evidence. Keyed on BOTH spellings,
# so a corrected source file makes the entry stop matching instead of silently
# applying a stale override.
#   (ags, hauptwahl-date, "Last, First" seen in HW, "Last, First" seen in SW)
#       -> canonical ("Last", "First")
NAME_FIXES = {
    # Kulike (HW) / Kulicke (SW). Local press and his own mayoral pages spell it
    # Kulicke (werneuchen-info.de, barnim-aktuell.de).
    ("12060280", "2019-09-01", "Kulike, Frank", "Kulicke, Frank"):
        ("Kulicke", "Frank"),
    # Heinz-Georg (HW) / Hein-Georg (SW). SPD Königs Wusterhausen and the town's
    # councillor register both give Heinz-Georg.
    ("12061260", "2017-09-24", "Hanke, Heinz-Georg", "Hanke, Hein-Georg"):
        ("Hanke", "Heinz-Georg"),
    # Alexander-Joachim (HW) / Alexander (SW). Both are his; the fuller form is
    # kept, as elsewhere in the dataset.
    ("12063056", "2013-04-14", "Lamprecht, Alexander-Joachim", "Lamprecht, Alexander"):
        ("Lamprecht", "Alexander-Joachim"),
    # Trebling (HW) / Tebling (SW). Premnitz's mayor is Ralf Tebling.
    ("12063244", "2016-04-10", "Trebling, Ralf", "Tebling, Ralf"):
        ("Tebling", "Ralf"),
    # Herinrich (HW) / Heinrich (SW). Gemeinde Glienicke/Nordbahn: Dr. Hans
    # Günther Heinrich Oberlack.
    ("12065096", "2017-09-24", "Oberlack, Hans Günther Herinrich",
     "Oberlack, Hans Günther Heinrich"): ("Oberlack", "Hans Günther Heinrich"),  # noqa: E501
    # Surname truncated to "K" in the Hauptwahl row. Vetschau/Spreewald's mayor
    # is Bengt Kanzler.
    ("12066320", "2017-09-24", "K, Bengt", "Kanzler, Bengt"): ("Kanzler", "Bengt"),
    # Name fields swapped in the Hauptwahl row ("Ronny, Kretschmer").
    ("12068320", "2013-01-13", "Ronny, Kretschmer", "Kretschmer, Ronny"):
        ("Kretschmer", "Ronny"),
    # Maass (HW) / Maaß (SW). Contemporary reporting (PNN) spells it Maaß.
    ("12069397", "2011-09-11", "Maass, Christian", "Maaß, Christian"):
        ("Maaß", "Christian"),
    # Fred (HW) / Fred Hans Willi (SW). The fuller form is kept.
    ("12071160", "2018-04-22", "Mahro, Fred", "Mahro, Fred Hans Willi"):
        ("Mahro", "Fred Hans Willi"),
}


def match_score(a, b):
    """Similarity of two 'Last, First' strings, tolerant of swapped fields."""
    a, b = a.casefold(), b.casefold()
    swap = lambda s: ", ".join(reversed([p.strip() for p in s.split(",", 1)])) \
        if "," in s else s
    return max(difflib.SequenceMatcher(None, a, b).ratio(),
               difflib.SequenceMatcher(None, swap(a), b).ratio())


def reconcile_rounds(deduped, problems):
    """Give one person one name across the Hauptwahl and Stichwahl of a cycle.

    Mutates the Stichwahl records in place (and applies NAME_FIXES to both
    rounds). Returns the number of candidates whose name was unified.
    """
    by_ags = defaultdict(list)
    for rec in deduped:
        by_ags[rec["_ags"]].append(rec)

    n_unified = 0
    for ags, recs in by_ags.items():
        hws = [r for r in recs if norm_ws(r["Wahlart"]) == "Hauptwahl"]
        sws = [r for r in recs if norm_ws(r["Wahlart"]) == "Stichwahl"]
        for sw in sws:
            prior = [h for h in hws
                     if 0 < (sw["_date"] - h["_date"]).days <= 60]
            if not prior:
                continue                       # orphan runoff, reported elsewhere
            hw = max(prior, key=lambda h: h["_date"])

            # Compare on the TITLE-STRIPPED "Last, First" form: "Dr. Radant,
            # Jana" and "Dr.phil. Radant, Jana" are the same record, and the
            # academic title is not part of the published name anyway.
            hw_names = [split_name(c["raw_name"])[0] for c in candidates_of(hw)]
            hw_parties = [c["party"] for c in candidates_of(hw)]

            # The source's own list of who reached the runoff, in HW spelling.
            quals = []
            for col in ("kandidat0_stichwahl", "kandidat1_stichwahl"):
                m = QUALIFIER_RE.match(norm_ws(hw.get(col)))
                if m:
                    quals.append(split_name(m.group(1))[0])
            pool = [q for q in quals if q in hw_names] or hw_names

            sw_c = candidates_of(sw)
            taken, assign = set(), {}
            # pass 1 — exact
            for i, c in enumerate(sw_c):
                nm = split_name(c["raw_name"])[0]
                if nm in pool and nm not in taken:
                    assign[i] = nm
                    taken.add(nm)
            # pass 2 — similarity, requiring a clear and unique best
            for i, c in enumerate(sw_c):
                if i in assign:
                    continue
                nm = split_name(c["raw_name"])[0]
                cand = [(match_score(nm, p), p) for p in pool if p not in taken]
                if not cand:
                    continue
                cand.sort(reverse=True)
                best, second = cand[0], (cand[1] if len(cand) > 1 else (0.0, None))
                same_party = (hw_parties[hw_names.index(cand[0][1])] == c["party"]
                              if cand[0][1] in hw_names else False)
                if best[0] >= 0.72 and best[0] - second[0] >= 0.08:
                    assign[i] = best[1]; taken.add(best[1])
                elif same_party and best[0] >= 0.55 and best[0] - second[0] >= 0.08:
                    assign[i] = best[1]; taken.add(best[1])
            # pass 3 — elimination: one runoff slot and one candidate left
            left_i = [i for i in range(len(sw_c)) if i not in assign]
            left_p = [p for p in pool if p not in taken]
            if len(left_i) == 1 and len(left_p) == 1:
                assign[left_i[0]] = left_p[0]; taken.add(left_p[0])

            for i, c in enumerate(sw_c):
                nm = split_name(c["raw_name"])[0]
                tgt = assign.get(i)
                if tgt is None:
                    problems["runoff_candidate_unmatched"].append(
                        f"{ags} {sw['_name']} {sw['_date']}: runoff candidate "
                        f"{nm!r} matches no Hauptwahl candidate")
                    continue
                if tgt == nm:
                    continue
                # Same person, two spellings. Unify so 01b pairs the rounds.
                key = (ags, hw["_date"].isoformat(), tgt, nm)
                fix = NAME_FIXES.get(key)
                canon = ("%s, %s" % fix) if fix else tgt
                problems["cross_round_name_variant"].append(
                    f"{ags} {sw['_name']} {hw['_date']}/{sw['_date']}: "
                    f"HW {tgt!r} vs SW {nm!r} -> {canon!r}"
                    f"{'' if fix else '  (no verified fix; kept the Hauptwahl spelling)'}")
                sw["_rename"] = sw.get("_rename", {})
                sw["_rename"][split_name(c["raw_name"])[0]] = canon
                if fix:
                    hw["_rename"] = hw.get("_rename", {})
                    hw["_rename"][tgt] = canon
                n_unified += 1
    return n_unified


def main():
    if not os.path.exists(XLSX):
        sys.exit(f"FATAL: source workbook not found: {XLSX}")

    raw = load_xlsx()
    print(f"Read {len(raw)} data rows from {os.path.basename(XLSX)}")

    problems = defaultdict(list)

    # ---- (c) rows without a usable ARS -------------------------------------
    keep = []
    for rec in raw:
        ags = ars_to_ags(rec["AGS/ARS"])
        if ags is None:
            problems["no_ags"].append(
                f"row {rec['_row']}: {norm_ws(rec['Gemeinde'])} "
                f"{rec['Datum der Wahl'].date()} [{norm_ws(rec['Wahlart'])}]")
            continue
        rec["_ags"] = ags
        # (d) strip the polling annotation from the Gemeinde name
        rec["_name"] = re.sub(r"\s+OT\s+.*$", "", norm_ws(rec["Gemeinde"])).strip()
        rec["_date"] = rec["Datum der Wahl"].date()
        keep.append(rec)

    # ---- (a)+(b) fingerprint dedup within an AGS ---------------------------
    groups = defaultdict(list)
    for rec in keep:
        groups[(rec["_ags"], fingerprint(rec))].append(rec)

    def richness(rec):
        """Prefer the copy that fills more Wahlvorschlagsart cells, then a
        Sunday date, then the later filing."""
        n_art = sum(1 for i in range(1, N_CAND_BLOCKS + 1)
                    if norm_ws(rec.get(f"art_{i}")))
        return (n_art, rec["Datum der Wahl"].weekday() == 6, rec["_date"])

    deduped = []
    for (ags, _fp), recs in groups.items():
        if len(recs) == 1:
            deduped.append(recs[0])
            continue
        best = max(recs, key=richness)
        for r in recs:
            if r is best:
                continue
            same_day = r["_date"] == best["_date"]
            problems["duplicate_same_day" if same_day else "duplicate_other_date"].append(
                f"{ags} {r['_name']}: dropped row {r['_row']} ({r['_date']}, "
                f"{r['Datum der Wahl'].strftime('%a')}) as a duplicate of row "
                f"{best['_row']} ({best['_date']})")
        deduped.append(best)
    deduped.sort(key=lambda r: (r["_ags"], r["_date"], r["Wahlart"]))

    # ---- give one person one name across the two rounds of a cycle ---------
    n_unified = reconcile_rounds(deduped, problems)
    if n_unified:
        print(f"  Cross-round name reconciliation: unified {n_unified} candidate "
              f"name(s) that the source spells differently in the two rounds")

    # ---- one canonical name per AGS ----------------------------------------
    name_votes = defaultdict(Counter)
    for rec in deduped:
        name_votes[rec["_ags"]][rec["_name"]] += 1
    canon_name = {a: c.most_common(1)[0][0] for a, c in name_votes.items()}

    # ---- build candidate-level rows ---------------------------------------
    out_rows = []
    n_sieger, n_sieger_ok = 0, 0
    for rec in deduped:
        ags = rec["_ags"]
        cands = candidates_of(rec)
        if not cands:
            problems["no_candidates"].append(
                f"row {rec['_row']}: {rec['_name']} {rec['_date']}")
            continue

        ev = rec["Wahlberechtigte"]
        nv = rec["Waehler"]
        valid = rec["gueltige Stimmen"]
        invalid = rec["ungueltige Stimmen"]

        # Arithmetic invariants of the source (all held at ingestion, 2026-08).
        if None not in (nv, valid, invalid) and valid + invalid != nv:
            problems["voters_ne_valid_plus_invalid"].append(
                f"row {rec['_row']}: {rec['_name']} {rec['_date']} "
                f"{valid}+{invalid} != {nv}")
        if ev and nv and nv > ev:
            problems["turnout_above_1"].append(
                f"row {rec['_row']}: {rec['_name']} {rec['_date']} {nv}/{ev}")
        vote_sum = sum(c["votes"] or 0 for c in cands)
        if valid:
            if len(cands) > 1 and vote_sum != valid:
                problems["votes_ne_valid"].append(
                    f"row {rec['_row']}: {rec['_name']} {rec['_date']} "
                    f"sum={vote_sum} valid={valid}")
            elif len(cands) == 1 and vote_sum > valid:
                problems["janein_above_valid"].append(
                    f"row {rec['_row']}: {rec['_name']} {rec['_date']} "
                    f"ja={vote_sum} valid={valid}")

        top = max(cands, key=lambda c: (c["votes"] is not None, c["votes"] or 0))
        n_top = sum(1 for c in cands if c["votes"] == top["votes"])
        if n_top > 1:
            problems["tied_lead"].append(
                f"row {rec['_row']}: {rec['_name']} {rec['_date']} "
                f"{n_top} candidates on {top['votes']} votes")

        # A single-candidate round is a Ja/Nein confirmation ballot: the Ja
        # votes are the candidate's, the Nein votes are the rest of `valid`.
        # Below 50 % the candidate is NOT elected — see the Schleswig-Holstein
        # lesson in CLAUDE.md; guard rather than seat a rejected candidate.
        lone_rejected = (len(cands) == 1 and valid and top["votes"] is not None
                         and top["votes"] / valid < 0.5)
        if lone_rejected:
            problems["janein_below_majority"].append(
                f"row {rec['_row']}: {rec['_name']} {rec['_date']} "
                f"{top['votes']}/{valid}")

        # Cross-check against the source's own "sieger" columns, which are
        # filled for the DECISIVE round from 2018 on (177 of 428 rounds).
        s_last = norm_ws(rec["sieger_nachname"])
        s_first = norm_ws(rec["sieger_vorname"])
        sieger_key = None
        if s_last:
            n_sieger += 1
            sl, _ = strip_title(s_last)
            sf, _ = strip_title(s_first)
            sieger_key = match_key(sl, sf)
            _, tl, tf, _ = split_name(rec.get("_rename", {}).get(
                split_name(top["raw_name"])[0], top["raw_name"]))
            if match_key(tl, tf) == sieger_key:
                n_sieger_ok += 1
            else:
                problems["sieger_ne_top"].append(
                    f"row {rec['_row']}: {rec['_name']} {rec['_date']} "
                    f"sieger={s_last}, {s_first} top={tl}, {tf}")

        etype = ("Oberbürgermeisterwahl" if ags in KREISFREIE
                 else "Bürgermeisterwahl")
        rnd = "stichwahl" if norm_ws(rec["Wahlart"]) == "Stichwahl" else "hauptwahl"
        turnout = (nv / ev) if (ev and nv) else ""

        renames = rec.get("_rename", {})
        for c in cands:
            # reconcile_rounds() may have unified this name with the other round
            raw_name = renames.get(split_name(c["raw_name"])[0], c["raw_name"])
            full, last, first, title = split_name(raw_name)
            title = title or c["title_col"]
            share = (c["votes"] / valid) if (valid and c["votes"] is not None) else ""
            is_winner = (c is top) and not lone_rejected
            out_rows.append({
                "ags": ags,
                "ags_name": canon_name[ags],
                "state": STATE, "state_name": STATE_NAME,
                "election_year": rec["_date"].year,
                "election_date": rec["_date"].isoformat(),
                "election_type": etype,
                "round": rnd,
                "eligible_voters": ev if ev is not None else "",
                "number_voters": nv if nv is not None else "",
                "valid_votes": valid if valid is not None else "",
                "invalid_votes": invalid if invalid is not None else "",
                "turnout": turnout,
                "candidate_name": full,
                "candidate_last_name": last,
                "candidate_first_name": first,
                "candidate_title": title,
                "candidate_party": c["party"],
                "wahlvorschlag_art": c["art"],
                "candidate_votes": c["votes"] if c["votes"] is not None else "",
                "candidate_voteshare": share,
                "is_winner": "TRUE" if is_winner else "FALSE",
                "source": "lwl_xlsx",
                "source_url": f"{rec['_sheet']}!row{rec['_row']}",
            })

    print(f"Parsed {len(out_rows)} candidate rows across "
          f"{len({(r['ags'], r['election_date'], r['round']) for r in out_rows})} rounds, "
          f"{len({r['ags'] for r in out_rows})} Gemeinden")
    if n_sieger:
        print(f"  Winner cross-check: {n_sieger_ok}/{n_sieger} rounds where the "
              f"source names the elected person agree with the top-voted "
              f"candidate of that round")

    # ---- merge the portal scrape for elections the XLSX does not carry -----
    n_portal = 0
    if os.path.exists(PORTAL_CSV):
        with open(PORTAL_CSV, encoding="utf-8") as fh:
            portal = list(csv.DictReader(fh))
        have = {(r["ags"], r["election_date"], r["round"]) for r in out_rows}
        extras = [r for r in portal
                  if (r["ags"], r["election_date"], r["round"]) not in have]

        # Same guard as the Sachsen-Anhalt merge: a portal round whose result
        # fingerprint already exists for that AGS under a different date is the
        # SAME election filed twice, not a missing one.
        def fp_rows(rows):
            g = defaultdict(list)
            for r in rows:
                g[(r["ags"], r["election_date"], r["round"])].append(r)
            out = {}
            for k, v in g.items():
                votes = sorted(int(x["candidate_votes"]) for x in v
                               if str(x["candidate_votes"]).strip().isdigit())
                out[k] = (k[2], str(v[0]["valid_votes"]).strip(), tuple(votes))
            return out

        xl_fp = {(k[0], f) for k, f in fp_rows(out_rows).items()}
        drop = set()
        for k, f in fp_rows(extras).items():
            if (k[0], f) in xl_fp:
                drop.add(k)
                problems["portal_dup_other_date"].append(
                    f"{k[0]} {k[1]} {k[2]}: same result as an XLSX round at "
                    f"another date")
        extras = [r for r in extras
                  if (r["ags"], r["election_date"], r["round"]) not in drop]

        by_round = defaultdict(list)
        for r in extras:
            by_round[(r["ags"], r["election_date"], r["round"])].append(r)
        for _k, rs in sorted(by_round.items()):
            votes = [int(x["candidate_votes"]) if str(x["candidate_votes"]).strip().isdigit()
                     else None for x in rs]
            mx = max((v for v in votes if v is not None), default=None)
            for r, v in zip(rs, votes):
                # The portal glues the academic title to the SURNAME ("Dr.
                # Strasser") where the XLSX puts it before the given name; strip
                # it the same way so one person has one name across sources.
                p_full, p_last, p_first, p_title = split_name(r["candidate_name"])
                # The portal prints the share rounded to 4 decimals; recompute
                # it so winner_voteshare == winner_votes / valid_votes holds for
                # every Brandenburg row, XLSX or portal.
                p_valid = str(r["valid_votes"]).strip()
                p_share = ((v / int(p_valid)) if (v is not None and p_valid.isdigit()
                                                  and int(p_valid) > 0)
                           else r["candidate_voteshare"])
                out_rows.append({
                    "ags": r["ags"], "ags_name": r["ags_name"],
                    "state": r["state"], "state_name": r["state_name"],
                    "election_year": r["election_year"],
                    "election_date": r["election_date"],
                    "election_type": r["election_type"], "round": r["round"],
                    "eligible_voters": r["eligible_voters"],
                    "number_voters": r["number_voters"],
                    "valid_votes": r["valid_votes"],
                    "invalid_votes": r["invalid_votes"],
                    "turnout": r["turnout"],
                    "candidate_name": p_full,
                    "candidate_last_name": p_last,
                    "candidate_first_name": p_first,
                    "candidate_title": p_title,
                    "candidate_party": r["candidate_party"],
                    "wahlvorschlag_art": "",
                    "candidate_votes": r["candidate_votes"],
                    "candidate_voteshare": p_share,
                    # Portal rows carry no elected-person record; leave the flag
                    # empty so the R stage falls back to rank(-votes) == 1.
                    "is_winner": "TRUE" if (v is not None and v == mx) else "",
                    "source": "portal",
                    "source_url": r["source_url"],
                })
                n_portal += 1
        print(f"  Portal fallback: {n_portal} candidate rows across "
              f"{len(by_round)} rounds the XLSX does not carry")
        for k in sorted(by_round):
            print(f"    + {k[0]} {by_round[k][0]['ags_name']} {k[1]} {k[2]}")
    else:
        print(f"  NOTE: portal scrape not found at {PORTAL_CSV} — XLSX only")

    # ---- structural checks over the merged set ----------------------------
    rounds = defaultdict(list)
    for r in out_rows:
        rounds[(r["ags"], r["election_date"], r["round"])].append(r)
    import datetime as _dt
    hw = defaultdict(list)
    sw = defaultdict(list)
    for (ags, d, rnd) in rounds:
        (sw if rnd == "stichwahl" else hw)[ags].append(_dt.date.fromisoformat(d))
    for ags, ds in sw.items():
        for d in ds:
            if not any(0 < (d - h).days <= 60 for h in hw.get(ags, [])):
                problems["orphan_stichwahl"].append(
                    f"{ags} {rounds[(ags, d.isoformat(), 'stichwahl')][0]['ags_name']} "
                    f"{d}: no Hauptwahl in the preceding 60 days")
    for key, rs in rounds.items():
        d = _dt.date.fromisoformat(key[1])
        if d.weekday() != 6:
            problems["non_sunday"].append(
                f"{key[0]} {rs[0]['ags_name']} {key[1]} ({d.strftime('%A')}) {key[2]}")

    # ---- report ------------------------------------------------------------
    print("\n--- source defects & anomalies ---")
    if not problems:
        print("  none")
    for k in sorted(problems):
        print(f"  [{k}] {len(problems[k])}")
        for msg in problems[k]:
            print(f"      {msg}")

    fatal = [k for k in ("votes_ne_valid", "voters_ne_valid_plus_invalid",
                         "janein_above_valid", "tied_lead", "sieger_ne_top")
             if problems.get(k)]
    if fatal:
        sys.exit(f"\nFATAL: unexpected source inconsistency in {fatal} — "
                 "investigate before publishing.")

    with open(OUT_CSV, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(out_rows[0].keys()))
        w.writeheader()
        w.writerows(out_rows)

    years = sorted({int(r["election_year"]) for r in out_rows})
    print(f"\nWrote {len(out_rows)} candidate rows -> {OUT_CSV}")
    print(f"  {len(rounds)} rounds, {len({r['ags'] for r in out_rows})} Gemeinden, "
          f"{years[0]}-{years[-1]}")
    print("  by election_type:",
          dict(Counter(r["election_type"] for r in out_rows)))
    print("  by round:", dict(Counter(r["round"] for r in out_rows)))


if __name__ == "__main__":
    main()
