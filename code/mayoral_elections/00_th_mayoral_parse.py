#!/usr/bin/env python3
"""
Stage 0 — Thüringen Oberbürgermeisterwahlen of the 6 kreisfreie Städte.

These OB elections are NOT in the wahlen.thueringen.de Bürgermeister (BM) database
(see 00_th_scrape.py) — they live in the combined "Landräte und Oberbürgermeister"
files under data/mayoral_elections/raw/thueringen/:

  * LRInfoG{YYYY}.xlsx / LSInfoG{YYYY}.xlsx  (2006, 2012, 2018, 2024)
      sheets "Kreis 051".."Kreis 056" = OB of the kreisfreie Städte.
      Candidate-level (name + party + votes). LR = Hauptwahl, LS = Stichwahl.
  * "Daten*.xls" (HTML tables, 1994 & 2000) — combined LR+OB, Gemeinde-level,
      party-level only (no candidate names), absolute counts, HW + SW.

The Landrat sheets (061..077) in the same xlsx are handled by the landrat pipeline
(code/landrat_elections/00_th_parse.R) — we take only the OB sheets here.

Cities (sheet 05N -> AGS 160 5N 000):
  051 Erfurt, 052 Gera, 053 Jena, 054 Suhl, 055 Weimar, 056 Eisenach.

Output: data/mayoral_elections/raw/thueringen/th_ob_parsed.csv (candidate-level long;
same columns as 00_th_scrape.py output for easy concatenation).
"""

import os
import re
import csv
import glob
import datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
RAW = os.path.join(ROOT, "data", "mayoral_elections", "raw", "thueringen")
OUT_CSV = os.path.join(RAW, "th_ob_parsed.csv")

CITY = {"051": "Erfurt", "052": "Gera", "053": "Jena",
        "054": "Suhl", "055": "Weimar", "056": "Eisenach"}

# Combined-election dates for the Daten HTML years (Hauptwahl / Stichwahl).
DATEN = {
    "1994": {"hw": ("Daten (5).xls", "1994-06-12"), "sw": ("Daten (6).xls", "1994-06-26")},
    "2000": {"hw": ("Daten (1).xls", "2000-05-14"), "sw": ("Daten (2).xls", "2000-05-28")},
}

NAME_RE = re.compile(r"^\s*[A-ZÄÖÜ][\wäöüÄÖÜß.\- ]+,\s+[A-ZÄÖÜ]")


def num(x):
    if x is None:
        return None
    s = re.sub(r"[^\d]", "", str(x))
    return int(s) if s else None


# ---------------------------------------------------------------------------
# Part A — Info xlsx OB sheets (candidate-level)
# ---------------------------------------------------------------------------
def parse_info_sheet(grid, sheet, round_):
    """grid = list of row-lists (cell values). Returns list of candidate dicts."""
    kd = str(re.search(r"\d+", sheet).group()).zfill(3)
    if kd not in CITY:
        return []
    title = " ".join(str(c) for c in grid[0] if c)
    if "bürgermeister" not in title.lower():
        return []
    year = int(re.search(r"(19|20)\d{2}", title).group()) if re.search(r"(19|20)\d{2}", title) else None

    # date from "Stand: DD.MM.YYYY" in first 4 rows
    edate = None
    for r in grid[:4]:
        for c in r:
            m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", str(c) if c else "")
            if m:
                edate = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
                break
        if edate:
            break

    # K-row: col2 (idx1) == "K" and col4 (idx3) == "000"
    k_idx = None
    for i, r in enumerate(grid):
        if len(r) > 3 and str(r[1]).strip() == "K" and str(r[3]).strip() == "000":
            k_idx = i
            break
    if k_idx is None:
        return []
    krow = grid[k_idx]

    # name-row: nearest row above K with "Last, First" cells; party-row = one above
    name_row = party_row = None
    cand_cols = []
    for tr in range(k_idx - 1, max(-1, k_idx - 7), -1):
        row = grid[tr]
        cols = [j for j, c in enumerate(row) if c and NAME_RE.match(str(c))]
        if cols:
            name_row = row
            cand_cols = cols
            party_row = grid[tr - 1] if tr - 1 >= 0 else None
            break
    if not cand_cols:
        return []

    elig = num(krow[8]) if len(krow) > 8 else None
    voters = num(krow[9]) if len(krow) > 9 else None
    invalid = num(krow[11]) if len(krow) > 11 else None
    valid = num(krow[12]) if len(krow) > 12 else None

    out = []
    for j in cand_cols:
        raw_name = str(name_row[j]).strip()
        pm = re.search(r"\(([^)]+)\)\s*$", raw_name)
        name = re.sub(r"\s*\([^)]+\)\s*$", "", raw_name).strip()
        party = pm.group(1).strip() if pm else None
        if not party and party_row is not None and j < len(party_row) and party_row[j]:
            party = str(party_row[j]).strip()
        votes = num(krow[j]) if j < len(krow) else None
        if votes is None:
            continue
        out.append(dict(
            ags="16" + kd + "000", ags_name=CITY[kd], election_year=year,
            election_date=edate, round=round_, election_type="Oberbürgermeisterwahl",
            amt="hauptamtlich", eligible_voters=elig, number_voters=voters,
            valid_votes=valid, invalid_votes=invalid,
            candidate_name=name, candidate_party=party, candidate_votes=votes,
            candidate_voteshare=round(votes / valid, 4) if valid else None,
            source="LAIV-MV Info xlsx" if False else "ThLfS Info xlsx",
        ))
    return out


def parse_info_files():
    rows = []
    for f in sorted(glob.glob(os.path.join(RAW, "L[RS]InfoG[0-9][0-9][0-9][0-9].xlsx"))):
        base = os.path.basename(f)
        if re.search(r"_\d", base):              # single-Kreis mid-term (Landrat only)
            continue
        round_ = "stichwahl" if base.startswith("LS") else "hauptwahl"
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            n = re.search(r"\d+", sheet)
            if not n or not (51 <= int(n.group()) <= 56):
                continue
            ws = wb[sheet]
            grid = [list(r) for r in ws.iter_rows(values_only=True)]
            rows.extend(parse_info_sheet(grid, sheet, round_))
        wb.close()
    return rows


# ---------------------------------------------------------------------------
# Part B — Daten HTML (party-level, 1994 & 2000)
# ---------------------------------------------------------------------------
def parse_daten(path, year, edate, round_):
    raw = open(path, encoding="iso-8859-1", errors="replace").read()
    thead = re.search(r"<thead>(.*?)</thead>", raw, re.S | re.I)
    hdr = [re.sub(r"\s+", "", re.sub(r"<[^>]+>", "", c)) for c in
           re.findall(r"<td[^>]*>(.*?)</td>", thead.group(1), re.S | re.I)] if thead else []
    # locate aggregate + party columns by header label
    def idx(label):
        for i, h in enumerate(hdr):
            if h.lower().startswith(label.lower()):
                return i
        return None
    i_elig = idx("Wahl-berechtigte") or idx("Wahlberechtigte")
    i_vot = idx("Wähler")
    i_inv = idx("Ungültige")
    i_val = idx("Gültige")
    party_start = (i_val + 1) if i_val is not None else None
    parties = hdr[party_start:] if party_start else []

    rows = []
    body = raw[thead.end():] if thead else raw
    for trm in re.finditer(r"<tr\b.*?(?=<tr\b|</table)", body, re.S | re.I):
        cells = re.findall(r"<td[^>]*>(.*?)(?=<td\b|<tr\b|</tr|</table)", trm.group(0), re.S | re.I)
        cells = [re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", c)).strip() for c in cells]
        if len(cells) < (party_start or 99):
            continue
        gemnr = re.sub(r"[^\d]", "", cells[1]) if len(cells) > 1 else ""
        if gemnr not in ("51000", "52000", "53000", "54000", "55000", "56000"):
            continue
        kd = "0" + gemnr[:2]
        elig = num(cells[i_elig]); voters = num(cells[i_vot])
        invalid = num(cells[i_inv]); valid = num(cells[i_val])
        for pi, pname in enumerate(parties):
            ci = party_start + pi
            if ci >= len(cells):
                continue
            v = cells[ci].strip()
            if v in ("", "-"):
                continue
            votes = num(v)
            if votes is None:
                continue
            rows.append(dict(
                ags="16" + kd + "000", ags_name=CITY.get(kd, ""), election_year=year,
                election_date=edate, round=round_, election_type="Oberbürgermeisterwahl",
                amt="hauptamtlich", eligible_voters=elig, number_voters=voters,
                valid_votes=valid, invalid_votes=invalid,
                candidate_name="", candidate_party=pname, candidate_votes=votes,
                candidate_voteshare=round(votes / valid, 4) if valid else None,
                source="ThLfS Daten HTML",
            ))
    return rows


def parse_daten_files():
    rows = []
    for year, spec in DATEN.items():
        for rk in ("hw", "sw"):
            fn, edate = spec[rk]
            p = os.path.join(RAW, fn)
            if os.path.exists(p):
                rows.extend(parse_daten(p, int(year), edate,
                                        "hauptwahl" if rk == "hw" else "stichwahl"))
    return rows


def main():
    info = parse_info_files()
    daten = parse_daten_files()
    rows = info + daten
    cols = ["ags", "ags_name", "state", "state_name", "election_year", "election_date",
            "election_type", "round", "amt", "eligible_voters", "number_voters",
            "valid_votes", "invalid_votes", "turnout", "candidate_name",
            "candidate_party", "candidate_votes", "candidate_voteshare",
            "winner_name_raw", "gemnr", "source"]
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            r.setdefault("state", "16"); r.setdefault("state_name", "Thüringen")
            t = (r["number_voters"] / r["eligible_voters"]) if (r.get("number_voters") and r.get("eligible_voters")) else ""
            r["turnout"] = round(t, 4) if t else ""
            r["winner_name_raw"] = ""
            r["gemnr"] = r["ags"][2:7]
            w.writerow({k: r.get(k, "") for k in cols})
    # summary
    import collections
    by = collections.Counter((r["election_year"], r["round"]) for r in rows)
    print(f"OB rows: {len(rows)} | Info(2006-2024): {len(info)} | Daten(1994/2000): {len(daten)}")
    print(f"distinct (city,year,round): {len({(r['ags'],r['election_year'],r['round']) for r in rows})}")
    for k in sorted(by): print("  ", k, by[k])
    print("->", OUT_CSV)


if __name__ == "__main__":
    main()
